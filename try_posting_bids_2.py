import streamlit as st
from openpyxl import load_workbook
import pandas as pd
import os, glob
from datetime import date

def get_most_recent_file(file_pattern='*', directory='.'):
    files = glob.glob(os.path.join(directory, file_pattern))

    if not files:
        return None

    def extract_date_from_filename(filepath):
        filename = os.path.basename(filepath)

        name_without_ext = os.path.splitext(filename)[0]

        date_match = re.search(r'\d{4}-\d{2}-\d{2}', filename)
        if date_match:
            try:
                return datetime.strptime(date_match.group(), '%Y-%m-%d')
            except ValueError:
                pass

        return datetime.min

    return max(files, key=extract_date_from_filename)

def create_multi_sheet_app_openpyxl(xlsx_file_path):
    """Create Streamlit pages for each sheet using openpyxl"""
    
    try:
        # Load workbook with openpyxl
        workbook = load_workbook(xlsx_file_path, data_only=True)
        sheet_names = workbook.sheetnames
        
        st.title("Multi-Sheet Data Viewer (OpenPyXL)")
        st.write(f"File: {os.path.basename(xlsx_file_path)}")
        
        # Create sidebar for sheet selection
        selected_sheet = st.sidebar.selectbox(
            "Select Sheet", 
            sheet_names,
            help="Choose which sheet to display"
        )
        
        # Display selected sheet
        if selected_sheet:
            st.header(f"Sheet: {selected_sheet}")
            
            # Get the worksheet
            worksheet = workbook[selected_sheet]
            
            # Get sheet dimensions
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # Display sheet info
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Rows", max_row)
            with col2:
                st.metric("Columns", max_col)
            with col3:
                st.metric("Sheet", f"{sheet_names.index(selected_sheet) + 1}/{len(sheet_names)}")
            
            # Convert worksheet to DataFrame for display
            df = worksheet_to_dataframe(worksheet)
            
            # Display the data
            st.dataframe(df, use_container_width=True)
            
            # Optional: Add download button for individual sheet
            csv = df.to_csv(index=False)
            st.download_button(
                label=f"Download {selected_sheet} as CSV",
                data=csv,
                file_name=f"{selected_sheet}.csv",
                mime="text/csv"
            )
    
    except Exception as e:
        st.error(f"Error reading Excel file: {str(e)}")
        return None

def worksheet_to_dataframe(worksheet):
    """Convert openpyxl worksheet to pandas DataFrame"""
    data = []
    
    # Get all values from the worksheet
    for row in worksheet.iter_rows(values_only=True):
        data.append(row)
    
    if not data:
        return pd.DataFrame()
    
    # Create DataFrame with first row as headers
    if len(data) > 1:
        df = pd.DataFrame(data[1:], columns=data[0])
    else:
        df = pd.DataFrame(data)
    
    return df

# Alternative: More detailed worksheet information
def create_detailed_sheet_viewer_openpyxl(xlsx_file_path):
    """Create detailed view with cell formatting and properties"""
    
    try:
        workbook = load_workbook(xlsx_file_path, data_only=False)  # Keep formulas
        sheet_names = workbook.sheetnames
        
        st.title("Detailed Excel Viewer (OpenPyXL)")
        
        selected_sheet = st.sidebar.selectbox("Select Sheet", sheet_names)
        
        if selected_sheet:
            worksheet = workbook[selected_sheet]
            
            # Show detailed sheet information
            st.header(f"Sheet: {selected_sheet}")
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Max Row", worksheet.max_row)
            with col2:
                st.metric("Max Column", worksheet.max_column)
            with col3:
                st.metric("Data Rows", len([row for row in worksheet.iter_rows() if any(cell.value for cell in row)]))
            with col4:
                st.metric("Has Formulas", sum(1 for row in worksheet.iter_rows() for cell in row if cell.data_type == 'f'))
            
            # Options for viewing
            view_options = st.multiselect(
                "View Options",
                ["Show Formulas", "Show Cell Comments", "Show Merged Cells", "Color Code by Data Type"],
                default=["Color Code by Data Type"]
            )
            
            # Convert to DataFrame with additional information
            df = advanced_worksheet_to_dataframe(worksheet, view_options)
            
            # Display the data
            st.dataframe(df, use_container_width=True)
            
            # Show additional sheet properties
            if st.expander("Sheet Properties"):
                st.write(f"**Sheet Title:** {worksheet.title}")
                st.write(f"**Sheet State:** {worksheet.sheet_state}")
                if hasattr(worksheet, 'sheet_properties'):
                    st.write(f"**Tab Color:** {worksheet.sheet_properties.tabColor}")
                
                # Show merged cell ranges
                if worksheet.merged_cells:
                    st.write("**Merged Cell Ranges:**")
                    for merged_range in worksheet.merged_cells.ranges:
                        st.write(f"- {merged_range}")

    except Exception as e:
        st.error(f"Error reading Excel file: {str(e)}")

def advanced_worksheet_to_dataframe(worksheet, view_options):
    """Convert worksheet to DataFrame with advanced options"""
    data = []
    
    for row in worksheet.iter_rows():
        row_data = []
        for cell in row:
            cell_value = cell.value
            
            # Show formulas if requested
            if "Show Formulas" in view_options and cell.data_type == 'f':
                cell_value = f"={cell.value}" if cell.value else cell.value
            
            # Add comments if requested
            if "Show Cell Comments" in view_options and cell.comment:
                cell_value = f"{cell_value} [Comment: {cell.comment.text}]"
            
            row_data.append(cell_value)
        data.append(row_data)
    
    if not data:
        return pd.DataFrame()
    
    # Create DataFrame
    if len(data) > 1:
        df = pd.DataFrame(data[1:], columns=data[0])
    else:
        df = pd.DataFrame(data)
    
    return df

# Tabbed version with openpyxl
def create_tabbed_sheet_viewer_openpyxl(xlsx_file_path):
    """Create tabs for each sheet using openpyxl"""
    
    try:
        workbook = load_workbook(xlsx_file_path, data_only=True)
        sheet_names = workbook.sheetnames
        
        st.title("Multi-Sheet Data Viewer (Tabbed - OpenPyXL)")
        
        # Create tabs
        tabs = st.tabs(sheet_names)
        
        for i, sheet_name in enumerate(sheet_names):
            with tabs[i]:
                worksheet = workbook[sheet_name]
                df = worksheet_to_dataframe(worksheet)
                
                st.write(f"**Rows:** {worksheet.max_row} | **Columns:** {worksheet.max_column}")
                st.dataframe(df, use_container_width=True)
                
                # Download button for each sheet
                csv = df.to_csv(index=False)
                st.download_button(
                    label=f"Download as CSV",
                    data=csv,
                    file_name=f"{sheet_name}.csv",
                    mime="text/csv",
                    key=f"download_{sheet_name}"
                )
    
    except Exception as e:
        st.error(f"Error reading Excel file: {str(e)}")

# Complete solution combining both problems
def main():
    st.set_page_config(
        page_title="Excel Sheet Viewer (OpenPyXL)",
        page_icon="📊",
        layout="wide"
    )
    
    # Step 1: Get the most recent xlsx file
    xlsx_file = get_most_recent_xlsx_file()
    
    if xlsx_file:
        st.success(f"Found file: {os.path.basename(xlsx_file)}")
        
        # Step 2: Choose viewing method
        view_method = st.radio(
            "Choose viewing method:",
            ["Basic View (Sidebar)", "Basic View (Tabs)", "Detailed View"],
            horizontal=True
        )
        
        if view_method == "Basic View (Sidebar)":
            create_multi_sheet_app_openpyxl(xlsx_file)
        elif view_method == "Basic View (Tabs)":
            create_tabbed_sheet_viewer_openpyxl(xlsx_file)
        else:
            create_detailed_sheet_viewer_openpyxl(xlsx_file)
    else:
        st.error("No xlsx files found with date format in filename")

def get_most_recent_xlsx_file(directory=".", file_pattern="*.xlsx"):
    """Find the most recent xlsx file based on date in filename"""
    import glob
    from datetime import datetime
    import re
    
    files = glob.glob(os.path.join(directory, file_pattern))
    
    if not files:
        return None
    
    def extract_date_from_filename(filepath):
        filename = os.path.basename(filepath)
        date_match = re.search(r'\d{4}-\d{2}-\d{2}', filename)
        if date_match:
            try:
                return datetime.strptime(date_match.group(), '%Y-%m-%d')
            except ValueError:
                pass
        return datetime.min
    
    return max(files, key=extract_date_from_filename)

if __name__ == "__main__":
    main()